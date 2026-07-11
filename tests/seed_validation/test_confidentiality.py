"""F3.5 (MI-20): confidentiality lint over the seed's own identifiers and
generated sample files (CLAUDE.md §7/§11/§20 -- the project's release
blocker). Needs no database: it lints the generators' static constants and
the checked-in seed/sample_files/* content directly, so it stays fast
(<2 min acceptance criterion, docs/tasks/F3.5.md) and runnable without a
Postgres instance.

Includes a "test of the test" (F3.5's own testing note): deliberately
planted real-shaped identifiers that must be caught, so a silently broken
lint can't pass CI by accident.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = REPO_ROOT / "backend"
for path in (REPO_ROOT, BACKEND_DIR):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from openpyxl import load_workbook  # noqa: E402

from seed.generators.catalog import FAMILIES, LINE_CODES, MACHINES  # noqa: E402
from seed.generators.users import ROLE_DEFINITIONS  # noqa: E402
from seed.validation.confidentiality import check_email, check_identifier  # noqa: E402

SAMPLE_FILES_DIR = REPO_ROOT / "seed" / "sample_files"

# Identifier-shaped columns worth linting in the sample files (measurement
# value columns are numbers, not identifiers -- nothing to lint there).
SAMPLE_FILE_IDENTIFIER_COLUMNS = {"part_number", "batch_lot", "machine_code", "operator_identifier"}


def _catalog_identifiers() -> list[str]:
    identifiers: list[str] = ["Plant Demo Norte", "MI-DEMO-ORG", "Assembly Area 1 (Demo)"]
    for family_code, family_name, parts in FAMILIES:
        identifiers.append(family_code)
        identifiers.append(family_name)
        for part_code, part_name in parts:
            identifiers.append(part_code)
            identifiers.append(part_name)
    identifiers.extend(LINE_CODES)
    for machine_code, _machine_type in MACHINES:
        identifiers.append(machine_code)
    return identifiers


def test_lint_accepts_every_catalog_identifier() -> None:
    for identifier in _catalog_identifiers():
        assert check_identifier(identifier) == [], f"false positive on fictitious identifier {identifier!r}"


def test_lint_accepts_every_seed_user_email() -> None:
    emails = [email for _role, _description, email, _display_name in ROLE_DEFINITIONS]
    assert emails, "expected at least one seed user email to check"
    for email in emails:
        assert check_email(email) == [], f"false positive on fictitious email {email!r}"


def test_lint_accepts_generated_csv_sample_files() -> None:
    csv_files = sorted(SAMPLE_FILES_DIR.glob("*.csv"))
    assert csv_files, "expected seed/sample_files/*.csv to exist (run seed/scripts/generate_sample_files.py)"

    for csv_path in csv_files:
        with csv_path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                for column in SAMPLE_FILE_IDENTIFIER_COLUMNS & set(row):
                    value = row[column]
                    assert check_identifier(value) == [], (
                        f"false positive in {csv_path.name}, column {column!r}: {value!r}"
                    )


def test_lint_accepts_generated_xlsx_sample_file() -> None:
    xlsx_files = sorted(SAMPLE_FILES_DIR.glob("*.xlsx"))
    assert xlsx_files, "expected seed/sample_files/*.xlsx to exist (run seed/scripts/generate_sample_files.py)"

    for xlsx_path in xlsx_files:
        workbook = load_workbook(xlsx_path, read_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                header = [str(cell) for cell in next(rows)]
                identifier_indexes = [
                    index for index, column in enumerate(header) if column in SAMPLE_FILE_IDENTIFIER_COLUMNS
                ]
                for row in rows:
                    for index in identifier_indexes:
                        value = str(row[index])
                        assert check_identifier(value) == [], (
                            f"false positive in {xlsx_path.name}/{sheet_name}, column {header[index]!r}: {value!r}"
                        )
        finally:
            workbook.close()


def test_lint_catches_bmw_style_grouped_part_number() -> None:
    # Structural shape of a real 11-digit BMW part number (2-2-1-3-3 grouping)
    # -- not a real number, just its shape -- must never slip through as a
    # fictitious MI-DEMO-* identifier.
    violations = check_identifier("51 11 7 123 456")
    assert violations, "lint failed to catch a BMW-shaped part number"


def test_lint_catches_vag_style_grouped_part_number() -> None:
    # Structural shape of a VW/Audi/Porsche/Skoda/Seat part number.
    violations = check_identifier("8E0-123-456-A")
    assert violations, "lint failed to catch a VAG-shaped part number"


def test_lint_catches_generic_long_digit_run() -> None:
    violations = check_identifier("123456789")
    assert violations, "lint failed to catch a generic long numeric identifier"


def test_lint_catches_real_oem_name_token() -> None:
    for planted in ["BMW Bracket Front Left", "Mercedes-Benz Caliper", "supplied by Bosch"]:
        violations = check_identifier(planted)
        assert violations, f"lint failed to catch a real OEM name token in {planted!r}"


def test_lint_catches_real_oem_plant_token() -> None:
    for planted in ["Plant Munich", "Werk Dingolfing Halle 3", "Spartanburg Line 2"]:
        violations = check_identifier(planted)
        assert violations, f"lint failed to catch a real OEM plant name in {planted!r}"


def test_lint_catches_non_local_email_domain() -> None:
    for planted in ["ana.garcia@bmw.com", "quality@onkaizen-demo.io", "test@gmail.com"]:
        violations = check_email(planted)
        assert violations, f"lint failed to catch a non-.local email in {planted!r}"


def test_lint_does_not_flag_fictitious_mi_demo_prefix_even_if_digit_heavy() -> None:
    # MI-DEMO-* is this project's own scheme (docs/seed-data-strategy.md) and
    # must stay exempt from the part-number shape checks that would
    # otherwise flag its digit suffix.
    assert check_identifier("MI-DEMO-1001") == []
