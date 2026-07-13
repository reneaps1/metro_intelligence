"""F4.5 (MI-25): the manual-upload connector -- CSV/XLSX, hardened parsing.

This is the demo's only import path, but it is itself just an
``ImportConnector`` implementation (CLAUDE.md §3): a future watched-folder or
PolyWorks connector plugs into the same ``validate``/``parse_rows`` contract
without the import service or DB schema changing.

Security posture (CLAUDE.md §5, docs/security/network-segmentation.md §3.1):
this is the most exposed surface in the system -- it parses attacker-
controlled bytes. Both connectors below only ever *read* structure (csv
module, openpyxl in read-only/data-only mode); neither evaluates formulas,
executes macros, or trusts the client-supplied filename/content-type beyond
picking which connector to try. ``validate()`` must run, and pass, before
``parse_rows()`` is ever called.
"""

from __future__ import annotations

import csv
import io
import zipfile
from collections.abc import Iterator

from openpyxl import load_workbook

from app.connectors.base import FileValidationError, ImportConnector, ParsedRow

MAX_XLSX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100

# Signatures of file formats that are not CSV/XLSX but are commonly used to
# smuggle an executable/archive past an extension-only check (a renamed
# .exe/.elf, or a zip that isn't actually an OOXML workbook).
_BINARY_SIGNATURES: tuple[bytes, ...] = (
    b"MZ",  # Windows PE (.exe/.dll)
    b"\x7fELF",  # Linux ELF
    b"\xca\xfe\xba\xbe",  # Mach-O / Java class
    b"%PDF-",
)

# OWASP CSV-injection trigger characters: a cell starting with one of these
# can be interpreted as a formula/command by spreadsheet software if the
# exported file is later reopened in Excel/Sheets -- rejected even though
# nothing in this backend ever evaluates cell content. "-" is deliberately
# *not* an unconditional trigger: measurement deviations are routinely
# negative numbers ("-0.029484"), so a bare leading minus only counts as
# suspicious when the rest of the cell isn't a plain number (a real
# injection payload like "-2+3+cmd|' /C calc'!A1" fails that check; "-0.03"
# passes it).
_UNCONDITIONAL_FORMULA_TRIGGERS = ("=", "+", "@", "\t", "\r")


def _looks_like_plain_negative_number(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def row_has_formula_injection(row: dict[str, str]) -> str | None:
    """Return the offending column name if any cell looks like a formula-
    injection payload, else None."""
    for column, value in row.items():
        if not value:
            continue
        if value.startswith(_UNCONDITIONAL_FORMULA_TRIGGERS):
            return column
        if value.startswith("-") and not _looks_like_plain_negative_number(value):
            return column
    return None


class CsvConnector(ImportConnector):
    def validate(self, content: bytes, filename: str) -> None:
        if not content:
            raise FileValidationError("File is empty.")
        if any(content.startswith(sig) for sig in _BINARY_SIGNATURES):
            raise FileValidationError("File content does not look like a valid CSV file.")
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise FileValidationError("File is not valid UTF-8 text.") from exc
        sample = text[:4096]
        printable = sum(1 for ch in sample if ch.isprintable() or ch in "\r\n\t")
        if sample and printable / len(sample) < 0.85:
            raise FileValidationError("File content does not look like a valid CSV file.")

        # Formula injection is treated as a crafted attack on the whole
        # file, not an isolated data-quality issue in one row -- reject the
        # upload outright rather than quarantining just that row.
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            row_data = {key: (value or "") for key, value in row.items() if key is not None}
            offending_column = row_has_formula_injection(row_data)
            if offending_column is not None:
                raise FileValidationError(
                    f"File rejected: possible formula/command injection in column '{offending_column}'."
                )

    def parse_rows(self, content: bytes) -> Iterator[ParsedRow]:
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        for row_number, row in enumerate(reader, start=1):
            data = {key: (value or "") for key, value in row.items() if key is not None}
            yield ParsedRow(row_number=row_number, data=data)


class XlsxConnector(ImportConnector):
    def validate(self, content: bytes, filename: str) -> None:
        if not content:
            raise FileValidationError("File is empty.")
        if not content.startswith(b"PK\x03\x04"):
            raise FileValidationError("File content does not look like a valid XLSX file.")
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile as exc:
            raise FileValidationError("File content does not look like a valid XLSX file.") from exc

        if "[Content_Types].xml" not in archive.namelist():
            raise FileValidationError("File content does not look like a valid XLSX file.")

        total_uncompressed = 0
        for info in archive.infolist():
            total_uncompressed += info.file_size
            if info.compress_size > 0 and info.file_size / info.compress_size > MAX_XLSX_COMPRESSION_RATIO:
                raise FileValidationError("File failed a safety check (compression ratio too high).")
        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise FileValidationError("File failed a safety check (uncompressed size too large).")

        for row in self.parse_rows(content):
            offending_column = row_has_formula_injection(row.data)
            if offending_column is not None:
                raise FileValidationError(
                    f"File rejected: possible formula/command injection in column '{offending_column}'."
                )

    def parse_rows(self, content: bytes) -> Iterator[ParsedRow]:
        # read_only streams rows instead of materializing the whole sheet;
        # data_only returns cached values, never formula source strings --
        # openpyxl never evaluates formulas in either mode.
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            row_number = 0
            for sheet in workbook.worksheets:
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue
                columns = [str(cell) if cell is not None else "" for cell in header]
                for values in rows_iter:
                    if values is None or all(v is None for v in values):
                        continue
                    row_number += 1
                    data = {
                        columns[i]: ("" if value is None else str(value))
                        for i, value in enumerate(values)
                        if i < len(columns) and columns[i]
                    }
                    yield ParsedRow(row_number=row_number, data=data)
        finally:
            workbook.close()


_CONNECTORS_BY_EXTENSION: dict[str, ImportConnector] = {
    ".csv": CsvConnector(),
    ".xlsx": XlsxConnector(),
}


def connector_for_filename(filename: str) -> ImportConnector:
    extension = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    connector = _CONNECTORS_BY_EXTENSION.get(extension)
    if connector is None:
        allowed = ", ".join(sorted(_CONNECTORS_BY_EXTENSION))
        raise FileValidationError(f"Unsupported file type '{extension or filename}'. Allowed: {allowed}.")
    return connector
